# -*- coding: utf-8 -*-
"""Compare_Excels.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1vFcsbMdgFUQeMANAKk_dws5REkfEjUXu
"""

#!pip install xlwings

from pathlib import Path  # Core Python Module

import pandas as pd  # pip install pandas openpyxl
import xlwings as xw  # pip install xlwings

initial_version = Path.cwd() / "source_file.xlsx"
updated_version = Path.cwd() / "destination_file.xlsx"

df_initial = pd.read_excel(initial_version)
print(df_initial.head(3))

df_update = pd.read_excel(updated_version)
print(df_update.head(3))

df_initial.shape

df_update.shape

df_initial.shape == df_update.shape

diff = df_update.compare(df_initial, align_axis=1)
diff
# self = updated_version
# other = initial_version

diff = df_update.compare(df_initial, align_axis=0)
diff

diff = df_update.compare(df_initial, keep_shape=True, keep_equal=False)
diff

diff = df_update.compare(df_initial, keep_shape=True, keep_equal=True)
diff

diff = df_update.compare(df_initial, keep_shape=True, keep_equal=True, align_axis=1, result_names=("Target", "Source"))
diff
diff.to_excel(Path.cwd() / "Difference.xlsx")

import openpyxl
from openpyxl.styles import PatternFill

def compare_and_highlight(file_path):
    # Load the Excel file
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Iterate through each row and compare "source" and "target" columns
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        source_value = row[3].value
        target_value = row[4].value

        # Set the default fill color to green
        fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # If source and target values are different, change the fill color to red
        if source_value != target_value:
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Apply fill color to the entire row
        for cell in row:
            cell.fill = fill

    # Save the changes
    wb.save(file_path)

# Example usage:
compare_and_highlight("/content/Difference.xlsx")

